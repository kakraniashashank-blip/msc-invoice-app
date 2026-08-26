import streamlit as st
import json
import google.generativeai as genai
from PIL import Image
from datetime import date

# -------------------------------------------------------------
# Configuration
# -------------------------------------------------------------
st.set_page_config(page_title="MSC Invoice Generator", page_icon="📑", layout="centered")

# Get API key from Streamlit secrets or environment
if "GEMINI_API_KEY" in st.secrets:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
else:
    api_key_input = st.sidebar.text_input("Enter Gemini API Key", type="password")
    if api_key_input:
        genai.configure(api_key=api_key_input)

# Master item descriptions lookup
MASTER_DESCRIPTIONS = {
    "09083BL": 'BRIGHT M.S. 1.31/32" (50MM) DIA. TOLERANCE -0.001"/-0.004" (En3B-BS970:1955)',
    "09118BL": 'BRIGHT M.S. 1/2" DIA. TOLERANCE ON DIA. -0.001"/-0.003". LENGTH: 18-20 FEET (En3B-BS970:1955)',
    "09119BL": 'BRIGHT M.S. 5/8" DIA. TOLERANCE ON DIA. -0.001"/-0.003". LENGTH: 18-20 FEET (En3B-BS970:1955)',
    "09304BL": 'BRIGHT M.S. 7/8" DIA. TOLERANCE ON DIA. -0.001"/-0.003". LENGTH: 11-15 FEET (En3B-BS970:1955)',
    "09113BL": 'BRIGHT M.S. 2" DIA. TOLERANCE ON DIA. +0.005"/+0.008". LENGTH: 18.5 FEET (En3B-BS970:1955)',
    "09114BL": 'BRIGHT M.S. 2.1/4" DIA. TOLERANCE ON DIA. +0.005"/+0.008". LENGTH: 18.5 FEET (En3B-BS970:1955)',
    "09283BL": 'BRIGHT M.S. 1.1/2" DIA. TOLERANCE ON DIA. -0.002"/-0.005". LENGTH: 18-20 FEET (En3B-BS970:1955)',
    "09822BL": 'BRIGHT CK45 STEEL - 2.1/4" DIA. GB TOL ON OD +0.014"/+0.018", HARDNESS 204 TO 249 BHN',
    "09158BL": 'BRIGHT M.S. GB 45MM DIA. TOLERANCE ON DIA.',
    "09110BL": 'BRIGHT M.S. 1" DIA. TOLERANCE -0.005"/-0.008". LENGTH: 8-12 FT (En3B-BS970:1955)',
    "09111BL": 'BRIGHT MS 1.1/4" DIA. TOLERANCE -0.002"/-0.005". LENGTH: 20-21 FEET (En3B-BS970:1955)',
    "09337BL": 'BRIGHT M.S. 3/4" DIA. TOLERANCE -0.001"/-0.003". LENGTH: 8-12 FT (En3B-BS970:1955)',
    "09828BL": 'BRIGHT CK-45, 1.5/8" GB+',
    "09816BL": '1.3/8" DIA BRIGHT CLASS IV STEEL (G.B) TOL ON DIA +0.008"/+0.012"',
    "09246BL": 'M.S. CHANNEL 3" X 1.1/2" (75X40MM), LENGTH: 20-22 FT',
    "09256BL": 'M.S. CHANNEL 4" X 2" (100X50MM)',
    "09427BL": 'M.S. ANGLE 3" X 3" X 1/4" (75X75X6 MM)'
}

HSN_CODES = {
    "09246BL": "721610", "09256BL": "721610", "09427BL": "721610"
}

CLIENT_DATABASE = {
    "Lagan Engineering Co. Ltd.": {
        "Address": "14 KYD Street, Kolkata - 700016",
        "GSTIN": "19AAACT9986F1ZP",
        "State": "West Bengal   Code: 19",
        "Delivery": "Lagan Engineering Co. Ltd., c/o Angus Jute Mills, Bhadeshwar, W.B."
    },
    "Birla Corporation Ltd., Unit Birla Jute Mill": {
        "Address": "9/1 R.N. Mukherjee Road, Kolkata 700001",
        "GSTIN": "19AABCB2075J1ZN",
        "State": "West Bengal   Code: 19",
        "Delivery": "Birla Jute Mill, P.O.Birlapur, 24 Parganas, West Bengal"
    }
}

def num_to_words(num):
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
             "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def two_digits(n):
        if n < 20: return units[n]
        return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")

    rupees = int(num)
    paise = int(round((num - rupees) * 100))
    crore = rupees // 10000000; rupees %= 10000000
    lakh = rupees // 100000; rupees %= 100000
    thousand = rupees // 1000; rupees %= 1000
    hundred = rupees // 100; rupees %= 100
    
    parts = []
    if crore > 0: parts.append(two_digits(crore) + " Crore")
    if lakh > 0: parts.append(two_digits(lakh) + " Lakh")
    if thousand > 0: parts.append(two_digits(thousand) + " Thousand")
    if hundred > 0: parts.append(two_digits(hundred) + " Hundred")
    if rupees > 0: parts.append(two_digits(rupees))
    
    words = " ".join(parts) + " Rupees"
    if paise > 0: words += f" and {two_digits(paise)} Paise"
    return words + " Only"

def extract_bill_details(image_file):
    model = genai.GenerativeModel("gemini-3.6-flash")
    prompt = """
    Extract all billing and item details from this bill/PO into a clean JSON structure:
    {
      "order_no": "P/2627/1552",
      "order_date": "07-08-2026",
      "delivery_charges": 16500,
      "items": [
        {
          "code": "09083BL",
          "pcs": 0,
          "qty": 560.0,
          "rate": 70.0
        }
      ]
    }
    Extract handwritten overrides if present. Return ONLY valid JSON.
    """
    
    # Send directly to Gemini without extra libraries
    if image_file.name.lower().endswith('.pdf'):
        payload = [prompt, {"mime_type": "application/pdf", "data": image_file.getvalue()}]
    else:
        img = Image.open(image_file)
        payload = [prompt, img]

    response = model.generate_content(payload)
    clean_text = response.text.replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(clean_text)
    except Exception:
        return {}

import io

def create_excel(doc_type, inv_no, inv_date, selected_client_name, client_info, order_no, order_date, edited_items, del_charges):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Define styles
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=16, color="003399")
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin = Side(border_style="thin", color="000000")
    box_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    no_border = Border()

    # Set column widths
    cols = [("A", 6), ("B", 18), ("C", 50), ("D", 12), ("E", 10), ("F", 15), ("G", 15), ("H", 20)]
    for col, width in cols:
        ws.column_dimensions[col].width = width

    # Row 1: Header
    ws.merge_cells("C1:F1")
    ws["C1"].value = doc_type
    ws["C1"].font = bold_font
    ws["C1"].alignment = center_align

    ws.merge_cells("G1:H1")
    ws["G1"].value = "Original for Buyer/ Seller"
    ws["G1"].alignment = right_align

    # Row 2-4: Company Info
    ws.merge_cells("A2:H2")
    ws["A2"].value = "MURLI STEEL CORPORATION"
    ws["A2"].font = title_font
    ws["A2"].alignment = center_align

    ws.merge_cells("A3:H3")
    ws["A3"].value = "9/12, Lal Bazar Street, Mercantile Building, 'B' Block, 1st Floor, Kolkata - 700001, India\nPhone: (033) 2210 1650 | Mobile: 9830242818 | Email: shradkakrania@gmail.com"
    ws["A3"].alignment = center_align
    ws.row_dimensions[3].height = 30

    ws.merge_cells("A4:H4")
    ws["A4"].value = "PAN: AKAPK4846L | GSTIN: 19AKAPK4846L1ZS"
    ws["A4"].font = bold_font
    ws["A4"].alignment = center_align

    # Info table
    ws.merge_cells("A5:D5")
    ws["A5"].value = "BILLED TO PARTY"
    ws.merge_cells("E5:H5")
    ws["E5"].value = "INVOICE DETAILS"
    for col in ["A", "E"]:
        ws[f"{col}5"].font = bold_font
        ws[f"{col}5"].fill = header_fill
        ws[f"{col}5"].alignment = center_align
        ws[f"{col}5"].border = box_border
    
    # Merge cells for better layout and add borders
    info_rows = [
        ("Name:", selected_client_name, "Invoice No.:", inv_no),
        ("Address:", client_info.get('Address', ''), "Invoice Date:", inv_date),
        ("GSTIN:", client_info.get('GSTIN', ''), "Terms:", doc_type),
        ("Order No.:", order_no, "Supply:", "West Bengal"),
        ("Order Date:", order_date, "", ""),
        ("State:", client_info.get('State', ''), "", "")
    ]
    
    row_idx = 6
    for row_data in info_rows:
        ws[f"A{row_idx}"] = row_data[0]
        ws.merge_cells(f"B{row_idx}:D{row_idx}")
        ws[f"B{row_idx}"] = row_data[1]
        if "Name:" in row_data[0] or "GSTIN" in row_data[0] or "Invoice No." in row_data[2] or "Invoice Date" in row_data[2]:
            ws[f"B{row_idx}"].font = bold_font
            ws[f"F{row_idx}"].font = bold_font
            
        ws[f"E{row_idx}"] = row_data[2]
        ws.merge_cells(f"F{row_idx}:H{row_idx}")
        ws[f"F{row_idx}"] = row_data[3]
        
        # Add outer borders for info section
        for c in range(1, 9):
            if c == 1 or c == 5:
                ws.cell(row=row_idx, column=c).border = Border(left=thin)
            elif c == 4 or c == 8:
                ws.cell(row=row_idx, column=c).border = Border(right=thin)
        row_idx += 1

    # Delivery
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws[f"A{row_idx}"] = f"Delivery At: {client_info.get('Delivery', '')}"
    ws.merge_cells(f"E{row_idx}:F{row_idx}")
    ws[f"E{row_idx}"] = "Transport: Lorry"
    ws.merge_cells(f"G{row_idx}:H{row_idx}")
    ws[f"G{row_idx}"] = "Vehicle No.:"
    
    for c in [1, 5, 7]:
        ws.cell(row=row_idx, column=c).border = Border(top=thin, bottom=thin)
    row_idx += 1

    # Items Header
    headers = ["SN", "Item Code", "Description", "HSN", "Pcs", "Qty", "Rate", "Value (INR)"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c_idx, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = box_border
    row_idx += 1
    
    start_item_row = row_idx
    # Items
    for idx, itm in enumerate(edited_items, 1):
        desc = MASTER_DESCRIPTIONS.get(itm['code'], itm['code'])
        hsn = HSN_CODES.get(itm['code'], "721550")
        
        ws.cell(row=row_idx, column=1, value=idx).alignment = center_align
        ws.cell(row=row_idx, column=2, value=itm['code']).alignment = center_align
        ws.cell(row=row_idx, column=3, value=desc).alignment = left_align
        ws.cell(row=row_idx, column=4, value=hsn).alignment = center_align
        ws.cell(row=row_idx, column=5, value=itm.get('pcs', '')).alignment = center_align
        ws.cell(row=row_idx, column=6, value=itm['qty']).alignment = center_align
        ws.cell(row=row_idx, column=7, value=itm['rate']).alignment = right_align
        
        # Formula for value
        ws.cell(row=row_idx, column=8, value=f"=F{row_idx}*G{row_idx}").alignment = right_align
        
        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).border = box_border
        row_idx += 1

    # Totals Row
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.cell(row=row_idx, column=1, value="TOTAL:").alignment = right_align
    ws.cell(row=row_idx, column=1).font = bold_font
    
    ws.cell(row=row_idx, column=6, value=f"=SUM(F{start_item_row}:F{row_idx-1})").font = bold_font
    ws.cell(row=row_idx, column=6).alignment = center_align
    ws.cell(row=row_idx, column=8, value=f"=SUM(H{start_item_row}:H{row_idx-1})").font = bold_font
    ws.cell(row=row_idx, column=8).alignment = right_align
    
    for c in range(1, 9):
        ws.cell(row=row_idx, column=c).border = box_border
        
    row_idx += 2
    
    # Totals Box
    ws.merge_cells(f"A{row_idx}:E{row_idx}")
    ws[f"A{row_idx}"] = "Total Invoice Amount in Words:"
    ws[f"A{row_idx}"].font = bold_font
    
    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = "Total Amount Before Tax"
    ws[f"H{row_idx}"] = f"=H{row_idx-2}"
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = box_border
    row_idx += 1

    # In Words string needs to be static since Excel formula for num2words is complex
    total_taxable = sum(it["qty"] * it["rate"] for it in edited_items)
    taxable_val = total_taxable + del_charges
    cgst = taxable_val * 0.09
    sgst = taxable_val * 0.09
    grand_total = taxable_val + cgst + sgst

    ws.merge_cells(f"A{row_idx}:E{row_idx+4}")
    ws[f"A{row_idx}"] = num_to_words(grand_total) + "\n(Note: Text will not auto-update if you edit values in Excel)"
    ws[f"A{row_idx}"].alignment = left_align
    
    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = "Delivery Charges"
    ws[f"H{row_idx}"] = del_charges
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = box_border
    row_idx += 1

    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = "Taxable Value"
    ws[f"H{row_idx}"] = f"=H{row_idx-2}+H{row_idx-1}"
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = box_border
    row_idx += 1

    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = "Add: CGST @ 9%"
    ws[f"H{row_idx}"] = f"=H{row_idx-1}*0.09"
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = box_border
    row_idx += 1

    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = "Add: SGST @ 9%"
    ws[f"H{row_idx}"] = f"=H{row_idx-2}*0.09"
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = box_border
    row_idx += 1

    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = "Grand Total"
    ws[f"F{row_idx}"].font = bold_font
    ws[f"F{row_idx}"].fill = header_fill
    ws[f"H{row_idx}"] = f"=H{row_idx-3}+H{row_idx-2}+H{row_idx-1}"
    ws[f"H{row_idx}"].font = bold_font
    ws[f"H{row_idx}"].fill = header_fill
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = box_border

    row_idx += 2
    # Footer
    ws.merge_cells(f"A{row_idx}:E{row_idx+3}")
    ws[f"A{row_idx}"] = "Bank Details :\nHDFC Bank Ltd. | A/c No.: 00082000057539\nBranch: Sree Bhumi | IFSC: HDFC0004566\nGoods once sold will not be taken back. E & O.E."
    ws[f"A{row_idx}"].alignment = left_align

    ws.merge_cells(f"F{row_idx}:H{row_idx+3}")
    ws[f"F{row_idx}"] = "Certified that the particulars given above are true and correct.\nFor MURLI STEEL CORPORATION\n\n\nAuthorised Signatory"
    ws[f"F{row_idx}"].alignment = right_align

    return wb

# -------------------------------------------------------------
# Streamlit App UI
# -------------------------------------------------------------
st.title("📄 Murli Steel Invoicer")

doc_type = st.radio("Document Type", ["PROFORMA INVOICE", "TAX INVOICE", "QUOTATION"], horizontal=True)

if doc_type == "PROFORMA INVOICE":
    default_inv = "PI/07/2026-27"
elif doc_type == "TAX INVOICE":
    default_inv = "MSC/20/2026-27"
else:
    default_inv = "QT/07/2026-27"

inv_no = st.text_input("Document Number", value=default_inv)
inv_date = st.text_input("Invoice Date", value=date.today().strftime("%d-%m-%Y"))

selected_client_name = st.selectbox("Select Client", list(CLIENT_DATABASE.keys()))
client_info = CLIENT_DATABASE[selected_client_name]

uploaded_file = st.file_uploader("📷 Snap Photo or Upload Bill", type=["jpg", "jpeg", "png", "pdf"])

if uploaded_file is not None:
    if "extracted_data" not in st.session_state or st.button("🔄 Re-Scan Image"):
        with st.spinner("Analyzing document..."):
            extracted = extract_bill_details(uploaded_file)
            if not extracted:
                st.error("Couldn't read the document clearly. Try a clearer image or manually enter details.")
                st.session_state.extracted_data = {}
            else:
                st.session_state.extracted_data = extracted

    data = st.session_state.extracted_data
    
    st.subheader("Step 1: Check Details")
    c1, c2 = st.columns(2)
    order_no = c1.text_input("Order No.", value=data.get("order_no", ""))
    order_date = c2.text_input("Order Date", value=data.get("order_date", ""))
    del_charges = st.number_input("Delivery Charges (₹)", value=float(data.get("delivery_charges", 0.0)))
    
    st.write("**Items List (Tap any box to adjust):**")
    edited_items = []
    for i, itm in enumerate(data.get("items", []), 1):
        with st.expander(f"Item #{i} - {itm.get('code', '')}", expanded=True):
            col_a, col_b, col_c, col_d = st.columns(4)
            code = col_a.text_input(f"Code #{i}", value=itm.get("code", ""), key=f"c_{i}")
            pcs = col_b.text_input(f"Pcs #{i}", value=str(itm.get("pcs", "")), key=f"p_{i}")
            qty = col_c.number_input(f"Qty (kg) #{i}", value=float(itm.get("qty", 0.0)), key=f"q_{i}")
            rate = col_d.number_input(f"Rate (₹/kg) #{i}", value=float(itm.get("rate", 0.0)), key=f"r_{i}")
            edited_items.append({"code": code, "qty": qty, "rate": rate, "pcs": pcs})

    # Calculations
    total_taxable = sum(it["qty"] * it["rate"] for it in edited_items)
    taxable_val = total_taxable + del_charges
    cgst = taxable_val * 0.09
    sgst = taxable_val * 0.09
    grand_total = taxable_val + cgst + sgst
    
    st.markdown("---")
    st.write(f"**Total Before Tax:** ₹{total_taxable:,.2f}")
    st.write(f"**Taxable Value:** ₹{taxable_val:,.2f}")
    st.write(f"### **Grand Total:** ₹{grand_total:,.2f}")

    btn_col1, btn_col2 = st.columns(2)
    
    with btn_col2:
        if st.button("📊 Generate Excel", use_container_width=True):
            wb = create_excel(doc_type, inv_no, inv_date, selected_client_name, client_info, order_no, order_date, edited_items, del_charges)
            excel_io = io.BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)
            st.download_button(
                label="📥 Download / Share Excel",
                data=excel_io,
                file_name=f"{inv_no.replace('/', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with btn_col1:
        if st.button("✅ Generate PDF", type="primary", use_container_width=True):
            try:
                from weasyprint import HTML
            except Exception as e:
                st.error("PDF generation is disabled on Windows. Please push to GitHub to use this feature on Streamlit Cloud.")
                st.stop()
                
            rows_html = ""
            for idx, itm in enumerate(edited_items, 1):
                desc = MASTER_DESCRIPTIONS.get(itm['code'], itm['code'])
                hsn = HSN_CODES.get(itm['code'], "721550")
                tax_val = itm['qty'] * itm['rate']
                rows_html += f"""
                <tr>
                    <td>{idx}</td>
                    <td>{itm['code']}</td>
                    <td style="text-align: left; font-size: 9px;">{desc}</td>
                    <td>{hsn}</td>
                    <td>{itm.get('pcs', '')}</td>
                    <td>{itm['qty']:,.2f}</td>
                    <td style="text-align: right;">{itm['rate']:,.2f}</td>
                    <td style="text-align: right;">{tax_val:,.2f}</td>
                </tr>
                """

            full_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
            <style>
                @page {{ size: A4; margin: 10mm; }}
                * {{ box-sizing: border-box; }}
                body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 0; padding: 0; color: #000; }}
                .invoice-box {{ border: 1px solid #000; width: 100%; }}
                .header {{ text-align: center; border-bottom: 1px solid #000; padding: 5px; }}
                .company-name {{ font-size: 24px; font-weight: bold; color: #003399; margin: 5px 0; }}
                .info-table {{ width: 100%; border-collapse: collapse; border-bottom: 1px solid #000; }}
                .info-table td {{ border: 1px solid #000; padding: 4px; vertical-align: top; }}
                .section-title {{ text-align: center; font-weight: bold; background-color: #f0f0f0; }}
                .items-table {{ width: 100%; border-collapse: collapse; }}
                .items-table th, .items-table td {{ border: 1px solid #000; padding: 4px; text-align: center; }}
                .items-table th {{ background-color: #f0f0f0; font-size: 10px; }}
                .min-height-row td {{ height: 100px; }}
                .totals-box {{ width: 40%; float: right; border-collapse: collapse; }}
                .totals-box td {{ border: 1px solid #000; padding: 4px; text-align: right; }}
                .footer-table {{ width: 100%; border-collapse: collapse; border-top: 1px solid #000; }}
                .footer-table td {{ padding: 4px; vertical-align: top; }}
                .clear {{ clear: both; }}
            </style>
            </head>
            <body>
            <div class="invoice-box">
                
                <div class="header">
                    <table style="width: 100%; border: none;">
                        <tr>
                            <td style="width:33%; border: none;"></td>
                            <td style="width:34%; text-align:center; border: none;">
                                <span style="border: 1px solid #000; padding: 2px 10px; font-weight: bold; font-size:12px;">{doc_type}</span>
                            </td>
                            <td style="width:33%; text-align:right; font-size:10px; border: none;">Original for Buyer/ Seller</td>
                        </tr>
                    </table>
                    <div class="company-name">MURLI STEEL CORPORATION</div>
                    <div style="font-size:11px;">9/12, Lal Bazar Street, Mercantile Building, 'B' Block, 1st Floor, Kolkata - 700001, India</div>
                    <div style="font-size:11px;">Phone: (033) 2210 1650 | Mobile: 9830242818 | Email: shradkakrania@gmail.com</div>
                    <div style="font-weight:bold; font-size:12px; margin-top:5px;">PAN: AKAPK4846L | GSTIN: 19AKAPK4846L1ZS</div>
                </div>

                <table class="info-table">
                    <tr>
                        <td style="width: 50%;" class="section-title">BILLED TO PARTY</td>
                        <td style="width: 50%;" class="section-title">INVOICE DETAILS</td>
                    </tr>
                    <tr>
                        <td style="padding:0;">
                            <table style="width:100%; border-collapse:collapse;">
                                <tr><td style="width:25%; border:none; padding:3px;">Name:</td><td style="border:none; padding:3px; font-weight:bold;">{selected_client_name}</td></tr>
                                <tr><td style="border:none; padding:3px;">Address:</td><td style="border:none; padding:3px;">{client_info['Address']}</td></tr>
                                <tr><td style="border:none; padding:3px;">GSTIN:</td><td style="border:none; padding:3px; font-weight:bold;">{client_info['GSTIN']}</td></tr>
                                <tr><td style="border:none; padding:3px;">Order No.:</td><td style="border:none; padding:3px;">{order_no}</td></tr>
                                <tr><td style="border:none; padding:3px;">Order Date:</td><td style="border:none; padding:3px;">{order_date}</td></tr>
                                <tr><td style="border:none; padding:3px;">State:</td><td style="border:none; padding:3px;">{client_info['State']}</td></tr>
                            </table>
                        </td>
                        <td style="padding:0;">
                            <table style="width:100%; border-collapse:collapse;">
                                <tr><td style="width:35%; border:none; padding:3px;">Invoice No.:</td><td style="border:none; padding:3px; font-weight:bold;">{inv_no}</td></tr>
                                <tr><td style="border:none; padding:3px;">Invoice Date:</td><td style="border:none; padding:3px; font-weight:bold;">{inv_date}</td></tr>
                                <tr><td style="border:none; padding:3px;">Terms:</td><td style="border:none; padding:3px;">{doc_type}</td></tr>
                                <tr><td style="border:none; padding:3px;">Supply:</td><td style="border:none; padding:3px;">West Bengal</td></tr>
                            </table>
                        </td>
                    </tr>
                </table>

                <table class="info-table" style="border-top:none;">
                    <tr>
                        <td style="width: 55%; border-top:none;">Delivery At: {client_info['Delivery']}</td>
                        <td style="width: 20%; border-top:none;">Transport: Lorry</td>
                        <td style="width: 25%; border-top:none;">Vehicle No. : </td>
                    </tr>
                </table>

                <table class="items-table" style="border-top:none;">
                    <tr>
                        <th style="width: 4%;">SN</th>
                        <th style="width: 14%;">Item Code</th>
                        <th style="width: 38%;">Description</th>
                        <th style="width: 8%;">HSN</th>
                        <th style="width: 6%;">Pcs</th>
                        <th style="width: 10%;">Qty</th>
                        <th style="width: 8%;">Rate</th>
                        <th style="width: 12%;">Value (INR)</th>
                    </tr>
                    {rows_html}
                    <tr class="min-height-row">
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                        <td style="border-bottom:none; border-top:none;"></td>
                    </tr>
                    <tr>
                        <td colspan="4" style="text-align: right; font-weight: bold;">TOTAL:</td>
                        <td></td>
                        <td style="font-weight: bold;">{sum(it['qty'] for it in edited_items):,.2f}</td>
                        <td></td>
                        <td style="font-weight: bold; text-align: right;">{total_taxable:,.2f}</td>
                    </tr>
                </table>

                <div style="width: 100%;">
                    <div style="width: 55%; float: left; padding: 10px;">
                        <div style="font-weight:bold;">Total Invoice Amount in Words:</div>
                        <div style="margin-top: 5px;">{num_to_words(grand_total)}</div>
                    </div>
                    <table class="totals-box">
                        <tr><td style="text-align: left;">Total Amount Before Tax</td><td style="width: 40%;">{total_taxable:,.2f}</td></tr>
                        <tr><td style="text-align: left;">Delivery Charges</td><td>{del_charges:,.2f}</td></tr>
                        <tr><td style="text-align: left;">Taxable Value</td><td>{taxable_val:,.2f}</td></tr>
                        <tr><td style="text-align: left;">Add: CGST @ 9%</td><td>{cgst:,.2f}</td></tr>
                        <tr><td style="text-align: left;">Add: SGST @ 9%</td><td>{sgst:,.2f}</td></tr>
                        <tr style="background-color: #f0f0f0;"><td style="text-align: left; font-weight:bold;">Grand Total</td><td style="font-weight:bold;">{grand_total:,.2f}</td></tr>
                    </table>
                    <div class="clear"></div>
                </div>

                <table class="footer-table">
                    <tr>
                        <td style="width: 50%; border-right: 1px solid #000; padding: 8px;">
                            <div style="font-weight: bold; margin-bottom: 5px;">Bank Details :</div>
                            <div>HDFC Bank Ltd. | A/c No.: 00082000057539</div>
                            <div>Branch: Sree Bhumi | IFSC: HDFC0004566</div>
                            <div style="margin-top: 15px; font-size: 9px;">Goods once sold will not be taken back. E & O.E.</div>
                        </td>
                        <td style="width: 50%; text-align: left; padding: 8px; padding-left: 20px;">
                            <div style="font-size: 10px;">Certified that the particulars given above are true and correct.</div>
                            <div style="font-weight: bold; margin-top: 10px;">For MURLI STEEL CORPORATION</div>
                            <div style="margin-top: 35px;">Authorised Signatory</div>
                        </td>
                    </tr>
                </table>
            </div>
            </body>
            </html>
            """
            
            pdf_bytes = HTML(string=full_html).write_pdf()
            st.download_button(
                label="📥 Download / Share PDF",
                data=pdf_bytes,
                file_name=f"{inv_no.replace('/', '_')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
